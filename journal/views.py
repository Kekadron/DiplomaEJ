from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
from django.db import IntegrityError
from datetime import datetime, timedelta
from .models import Lesson, Grade, Discipline
from students.models import Teacher, Group, Student, Institution
from datetime import date as dt_date, timedelta
from datetime import datetime
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
import pandas as pd

@login_required
def teacher_dashboard(request, date=None):
    if request.user.is_superuser or request.user.is_staff:
        return redirect('admin_dashboard')

    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, "У вас нет прав преподавателя.")
        return redirect('login')

    # Текущая дата
    if date:
        current_date = dt_date.fromisoformat(date)
    else:
        current_date = dt_date.today()

    previous_date = current_date - timedelta(days=1)
    next_date = current_date + timedelta(days=1)

    lessons = Lesson.objects.filter(
        teacher=teacher,
        date=current_date
    ).select_related('discipline', 'group')\
     .order_by('pair_number')

    context = {
        'teacher': teacher,
        'lessons': lessons,
        'date': current_date,
        'previous_date': previous_date,
        'next_date': next_date,
    }

    return render(request, 'journal/teacher_dashboard.html', context)

@login_required
def lesson_grades(request, lesson_id):
    """Просмотр и редактирование журнала оценок по занятию"""
    lesson = get_object_or_404(Lesson, id=lesson_id)

    # Проверка прав доступа
    is_owner = hasattr(request.user, 'teacher') and request.user.teacher == lesson.teacher
    if not (request.user.is_superuser or request.user.is_staff or is_owner):
        messages.error(request, "У вас нет доступа к этому занятию.")
        return redirect('teacher_dashboard')

    # === АВТОМАТИЧЕСКОЕ СОЗДАНИЕ ОЦЕНОК ДЛЯ НОВЫХ СТУДЕНТОВ ===
    students = lesson.group.students.all()
    for student in students:
        Grade.objects.get_or_create(
            student=student,
            lesson=lesson
        )
    # ========================================================

    # Получаем все оценки (уже с новыми студентами)
    grades = Grade.objects.filter(lesson=lesson).select_related('student').order_by(
        'student__last_name', 'student__first_name'
    )

    if request.method == 'POST':
        saved = 0
        for grade in grades:
            value = request.POST.get(f'grade_{grade.id}', '')
            comment = request.POST.get(f'comment_{grade.id}', '')

            if grade.value != value or grade.comment != comment:
                grade.value = value
                grade.comment = comment
                grade.save()
                saved += 1

        if saved > 0:
            messages.success(request, f'Сохранено {saved} изменений!')
        else:
            messages.info(request, 'Изменений не было')

        return redirect('lesson_grades', lesson_id=lesson.id)

    context = {
        'lesson': lesson,
        'grades': grades,
    }
    return render(request, 'journal/lesson_grades.html', context)

@login_required
def admin_dashboard(request):
    """Панель администратора / Завуча"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "Доступ запрещён.")
        return redirect('login')
    
    # Инициализируем заранее
    teacher_groups = Group.objects.none()
    teacher_disciplines = Discipline.objects.none()

    if request.user.is_superuser:
        total_institutions = Institution.objects.count()
        total_groups = Group.objects.count()
        total_students = Student.objects.count()
        total_teachers = Teacher.objects.count()
        total_disciplines = Discipline.objects.count()
        
        teacher_groups = Group.objects.all()
        teacher_disciplines = Discipline.objects.all()
        
        institution = Institution.objects.first()
        recent_lessons = Lesson.objects.select_related('group', 'discipline', 'teacher').order_by('-date', '-pair_number')[:10]
    else:
        teacher = Teacher.objects.get(user=request.user)
        institution = teacher.institution
        total_institutions = 1
        total_groups = Group.objects.filter(institution=institution).count()
        total_students = Student.objects.filter(group__institution=institution).count()
        total_teachers = Teacher.objects.filter(institution=institution).count()
        total_disciplines = Discipline.objects.filter(institution=institution).count()
        
        teacher_groups = Group.objects.filter(institution=institution)
        teacher_disciplines = Discipline.objects.filter(institution=institution)
        
        recent_lessons = Lesson.objects.filter(
            group__institution=institution
        ).select_related('group', 'discipline', 'teacher').order_by('-date', '-pair_number')[:10]

    context = {
        'total_institutions': total_institutions,
        'total_groups': total_groups,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_disciplines': total_disciplines,
        'teacher_groups': teacher_groups,
        'teacher_disciplines': teacher_disciplines,
        'recent_lessons': recent_lessons,
        'institution': institution,
        'is_superuser': request.user.is_superuser,
    }
    return render(request, 'journal/admin_dashboard.html', context)

@login_required
def group_list(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('login')

    if request.user.is_superuser:
        groups = Group.objects.select_related('institution').all()
    else:
        teacher = Teacher.objects.get(user=request.user)
        groups = Group.objects.filter(institution=teacher.institution).select_related('institution')

    context = {'groups': groups}
    return render(request, 'journal/group_list.html', context)

@login_required
def group_create(request):
    """Создание новой группы"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    if request.method == 'POST':
        name = request.POST.get('name')
        specialty = request.POST.get('specialty')
        start_year = request.POST.get('start_year')
        
        # Определяем institution_id
        if request.user.is_superuser:
            institution_id = request.POST.get('institution')
        else:
            try:
                teacher = Teacher.objects.get(user=request.user)
                institution_id = teacher.institution.id
            except Teacher.DoesNotExist:
                messages.error(request, 'У вас не привязано учебное заведение')
                return redirect('group_list')

        # Проверка на пустые значения
        if not name or not specialty or not start_year:
            messages.error(request, 'Заполните все поля')
            return redirect('group_create') 
        
        if not institution_id:
            messages.error(request, 'Не указано учебное заведение')
            return redirect('group_create')

        Group.objects.create(
            institution_id=institution_id,
            name=name,
            specialty=specialty,
            start_year=start_year
        )
        messages.success(request, 'Группа успешно создана!')
        return redirect('group_list')

    institutions = Institution.objects.all()
    context = {
        'institutions': institutions,
        'is_superuser': request.user.is_superuser,
    }
    
    if not request.user.is_superuser:
        try:
            teacher = Teacher.objects.get(user=request.user)
            context['teacher'] = teacher
        except Teacher.DoesNotExist:
            pass
    
    return render(request, 'journal/group_create.html', context)

@login_required
def group_edit(request, pk):
    if not request.user.is_staff:
        return redirect('student_schedule')

    group = Group.objects.get(id=pk)
    institutions = Institution.objects.all()

    if request.method == "POST":
        group.name = request.POST.get("name")
        group.specialty = request.POST.get("specialty")
        group.start_year = request.POST.get("start_year")
        group.institution_id = request.POST.get("institution")
        group.save()
        return redirect('group_list')

    return render(request, "journal/group_edit.html", {
        "group": group,
        "institutions": institutions
    })

@login_required
def group_delete(request, pk):
    Group.objects.get(id=pk).delete()
    return redirect('group_list')

def home_redirect(request):
    """Самая простая версия для диагностики"""
    if not request.user.is_authenticated:
        return redirect('login')

    # Прямые проверки
    if request.user.is_superuser or request.user.is_staff:
        return redirect('admin_dashboard')

    if hasattr(request.user, 'teacher'):
        return redirect('teacher_dashboard_today')

    # Проверка студента
    try:
        student = Student.objects.get(user=request.user)
        return redirect('student_dashboard')
    except:
        pass

    # Если ничего не подошло
    messages.error(request, f"Роль не определена для пользователя {request.user.username}")
    return redirect('login')

@login_required
def student_create(request):
    """Создание нового студента"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    if request.method == 'POST':
        group_id = request.POST.get('group')
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        middle_name = request.POST.get('middle_name', '')
        student_id = request.POST.get('student_id')

        Student.objects.create(
            group_id=group_id,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
            student_id=student_id
        )
        messages.success(request, 'Студент успешно создан!')
        return redirect('admin_dashboard')

    if request.user.is_superuser:
        # Суперпользователь видит все группы
        groups = Group.objects.all()
    else:
        teacher = Teacher.objects.get(user=request.user)
        groups = Group.objects.filter(institution=teacher.institution)
    context = {'groups': groups}
    return render(request, 'journal/student_create.html', context)

@login_required
def student_edit(request, student_id):
    """Редактирование данных студента"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    student = get_object_or_404(Student, id=student_id)

    if request.method == 'POST':
        student.last_name = request.POST.get('last_name')
        student.first_name = request.POST.get('first_name')
        student.middle_name = request.POST.get('middle_name', '')
        student.student_id = request.POST.get('student_id')
        student.group_id = request.POST.get('group')
        student.save()
        messages.success(request, 'Данные студента успешно обновлены!')
        return redirect('admin_dashboard')

    if request.user.is_superuser:
        groups = Group.objects.all()
    else:
        teacher = Teacher.objects.get(user=request.user)
        groups = Group.objects.filter(institution=teacher.institution)
    
    context = {
        'student': student,
        'groups': groups,
    }
    return render(request, 'journal/student_edit.html', context)

@login_required
def student_delete(request, student_id):
    """Удаление студента"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    student = get_object_or_404(Student, id=student_id)
    student.delete()
    messages.success(request, 'Студент успешно удалён!')
    return redirect('admin_dashboard')

@login_required
def teacher_create(request):
    """Создание нового преподавателя"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    if request.method == 'POST':
        # Создаём пользователя Django
        username = request.POST.get('username')
        password = request.POST.get('password')
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        middle_name = request.POST.get('middle_name', '')
        phone = request.POST.get('phone', '')
        institution_id = request.POST.get('institution')

        user = User.objects.create_user(username=username, password=password)
        user.is_staff = False  # по умолчанию — не админ
        user.save()

        Teacher.objects.create(
            user=user,
            institution_id=institution_id,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
            phone=phone
        )
        messages.success(request, 'Преподаватель успешно создан!')
        return redirect('admin_dashboard')

    if request.user.is_superuser:
        # Суперпользователь видит все группы
        institutions = Institution.objects.all()
    else:
        teacher = Teacher.objects.get(user=request.user)
        institutions = Institution.objects.filter(id=teacher.institution.id)
    context = {'institutions': institutions}
    return render(request, 'journal/teacher_create.html', context)

@login_required
def teacher_edit(request, teacher_id):
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    teacher = Teacher.objects.get(id=teacher_id)

    if request.method == 'POST':
        teacher.last_name = request.POST.get('last_name')
        teacher.first_name = request.POST.get('first_name')
        teacher.middle_name = request.POST.get('middle_name', '')
        teacher.phone = request.POST.get('phone', '')
        teacher.save()
        messages.success(request, 'Данные преподавателя успешно обновлены!')
        return redirect('teacher_list')

    context = {
        'teacher': teacher,
    }
    return render(request, 'journal/teacher_edit.html', context)

@login_required
def teacher_delete(request, teacher_id):
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    teacher = Teacher.objects.get(id=teacher_id)
    user = teacher.user
    teacher.delete()
    user.delete()  # удаляем связанный аккаунт
    messages.success(request, 'Преподаватель успешно удалён!')
    return redirect('teacher_list')

@staff_member_required
def protected_admin(request):
    return redirect('/admin/')

@login_required
def student_list(request):
    """Список студентов"""
    if request.user.is_superuser:
        students = Student.objects.select_related('group__institution').all()
    else:
        # Завуч видит только студентов своего заведения
        teacher = Teacher.objects.get(user=request.user)
        students = Student.objects.filter(group__institution=teacher.institution).select_related('group__institution')

    context = {'students': students}
    return render(request, 'journal/student_list.html', context)

@login_required
def teacher_list(request):
    """Список преподавателей"""
    if request.user.is_superuser:
        teachers = Teacher.objects.select_related('institution').all()
    else:
        # Завуч видит только преподавателей своего заведения
        teacher = Teacher.objects.get(user=request.user)
        teachers = Teacher.objects.filter(institution=teacher.institution).select_related('institution')

    context = {'teachers': teachers}
    return render(request, 'journal/teacher_list.html', context)

@login_required
def student_dashboard(request):
    """Личный кабинет студента"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "У вас нет прав студента.")
        return redirect('login')

    # Группируем оценки по дисциплине
    from collections import defaultdict
    grades_by_discipline = defaultdict(list)
    avg_scores = {}

    grades = Grade.objects.filter(student=student).select_related('lesson__discipline').order_by('lesson__date')

    for grade in grades:
        discipline_name = grade.lesson.discipline.name
        grades_by_discipline[discipline_name].append(grade)

    # Считаем средний по каждой дисциплине + общий
    overall_total = 0
    overall_count = 0

    for discipline_name, grades_list in grades_by_discipline.items():
        total = 0
        count = 0
        for grade in grades_list:
            if grade.value.isdigit() and grade.value in ['2', '3', '4', '5']:
                total += int(grade.value)
                count += 1
        
        if count > 0:
            avg = round(total / count, 2)
        else:
            avg = 0
        
        avg_scores[discipline_name] = avg
        overall_total += total
        overall_count += count

    # Общий средний балл
    overall_avg = round(overall_total / overall_count, 2) if overall_count > 0 else 0

    context = {
        'student': student,
        'grades_by_discipline': dict(grades_by_discipline),
        'avg_scores': avg_scores,
        'overall_avg': overall_avg,
        'overall_count': overall_count,
        'disciplines_count': len(grades_by_discipline),
    }
    return render(request, 'journal/student_dashboard.html', context)

@login_required
def schedule_create(request):
    """Завуч создаёт занятие для преподавателей своего заведения"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "Доступ запрещён.")
        return redirect('login')

    # Определяем заведение завуча
    if request.user.is_superuser:
        institutions = Institution.objects.all()
    else:
        teacher = Teacher.objects.get(user=request.user)
        institutions = Institution.objects.filter(id=teacher.institution.id)

    if request.method == 'POST':
        institution_id = request.POST.get('institution')
        discipline_id = request.POST.get('discipline')
        group_id = request.POST.get('group')
        teacher_id = request.POST.get('teacher')
        date = request.POST.get('date')
        topic = request.POST.get('topic', '')

        lesson = Lesson.objects.create(
            discipline_id=discipline_id,
            group_id=group_id,
            teacher_id=teacher_id,
            date=date,
            pair_number = request.POST.get('pair_number'),
            topic=topic
        )

        # Автоматически создаём записи Grade для всех студентов группы
        students = Group.objects.get(id=group_id).students.all()
        for student in students:
            Grade.objects.get_or_create(student=student, lesson=lesson)

        messages.success(request, f'Занятие "{lesson.discipline}" на {lesson.date} успешно создано!')
        return redirect('admin_dashboard')

    # Формируем данные для формы
    if request.user.is_superuser:
        disciplines = Discipline.objects.all()
    else:
        disciplines = Discipline.objects.filter(institution__in=institutions)
    groups = Group.objects.filter(institution__in=institutions)
    teachers = Teacher.objects.filter(
        institution__in=institutions,
        user__is_staff=False  # исключаем сотрудников (завучей, админов)
    )

    context = {
        'disciplines': disciplines,
        'groups': groups,
        'teachers': teachers,
        'institutions': institutions,
    }
    return render(request, 'journal/schedule_create.html', context)

@login_required
def schedule_edit(request, lesson_id):
    """Редактирование занятия"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "Доступ запрещён.")
        return redirect('login')

    lesson = get_object_or_404(Lesson, id=lesson_id)

    # Определяем заведение завуча
    if request.user.is_superuser:
        institutions = Institution.objects.all()
    else:
        teacher = Teacher.objects.get(user=request.user)
        institutions = Institution.objects.filter(id=teacher.institution.id)

    if request.method == 'POST':
        lesson.discipline_id = request.POST.get('discipline')
        lesson.group_id = request.POST.get('group')
        lesson.teacher_id = request.POST.get('teacher')
        lesson.date = request.POST.get('date')
        lesson.pair_number = request.POST.get('pair_number')
        lesson.topic = request.POST.get('topic', '')
        lesson.save()

        messages.success(request, f'Занятие "{lesson.discipline}" на {lesson.date} успешно обновлено!')
        return redirect('admin_dashboard')

    # Формируем данные для формы
    if request.user.is_superuser:
        disciplines = Discipline.objects.all()
    else:
        disciplines = Discipline.objects.filter(institution__in=institutions)
    groups = Group.objects.filter(institution__in=institutions)
    teachers = Teacher.objects.filter(
        institution__in=institutions,
        user__is_staff=False  # исключаем сотрудников (завучей, админов)
    )

    context = {
        'lesson': lesson,
        'disciplines': disciplines,
        'groups': groups,
        'teachers': teachers,
        'institutions': institutions,
    }
    return render(request, 'journal/schedule_edit.html', context)

@login_required
def student_schedule(request, date=None):
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "У вас нет прав студента.")
        return redirect('login')

    # Если дата не передана → берём сегодня
    if date:
        try:
            current_date = datetime.strptime(date, "%Y-%m-%d").date()
        except:
            current_date = datetime.today().date()
    else:
        current_date = datetime.today().date()

    # Предыдущий и следующий день
    previous_date = (current_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (current_date + timedelta(days=1)).strftime("%Y-%m-%d")

    # Фильтрация занятий по дате
    lessons = Lesson.objects.filter(
        group=student.group,
        date=current_date
    ).select_related('discipline', 'teacher').order_by('pair_number')

    context = {
        'student': student,
        'lessons': lessons,
        'date': current_date,
        'previous_date': previous_date,
        'next_date': next_date,
    }

    return render(request, 'journal/student_schedule.html', context)

@login_required
def import_data(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('login')

    if request.method == 'POST':
        import_type = request.POST.get('import_type')
        file = request.FILES.get('file')

        if not file:
            messages.error(request, "Файл не выбран")
            return redirect('admin_dashboard')

        if not import_type:
            messages.error(request, "Тип импорта не выбран")
            return redirect('admin_dashboard')

        try:
            df = pd.read_excel(file)
            
            handlers = {
                'schedule': import_schedule_data,
                'students': import_students_data,
                'teachers': import_teachers_data,
                'disciplines': import_disciplines_data,
                'groups': import_groups_data,
            }
            
            handler = handlers.get(import_type)
            if handler:
                handler(request, df)
            else:
                messages.error(request, "Неизвестный тип импорта")

        except Exception as e:
            messages.error(request, f"Ошибка при чтении файла: {e}")

        return redirect('admin_dashboard')

    return redirect('admin_dashboard')

def import_schedule_data(request, df):
    created = 0
    skipped = 0
    errors = 0
    conflicts = []
    error_details = []
    
    institution = Institution.objects.first()
    if not institution:
        messages.error(request, "❌ Нет учебного заведения. Сначала создайте учреждение в админ-панели.")
        return
    
    # Кешируем справочники
    group_names = df['Группа'].dropna().unique()
    discipline_names = df['Дисциплина'].dropna().unique()
    
    groups_cache = {
        g.name: g for g in Group.objects.filter(name__in=group_names, institution=institution)
    }
    disciplines_cache = {
        d.name: d for d in Discipline.objects.filter(name__in=discipline_names, institution=institution)
    }
    
    teacher_last_names = df['Преподаватель'].dropna().apply(
        lambda x: str(x).split()[0]
    ).unique()
    teachers_cache = {
        t.last_name: t for t in Teacher.objects.filter(
            last_name__in=teacher_last_names, 
            institution=institution
        )
    }
    
    # Проверяем не найденные группы и дисциплины
    missing_groups = set(group_names) - set(groups_cache.keys())
    missing_disciplines = set(discipline_names) - set(disciplines_cache.keys())
    
    if missing_groups:
        messages.warning(request, f"⚠️ Группы не найдены в БД: {', '.join(sorted(missing_groups)[:5])}")
    if missing_disciplines:
        messages.warning(request, f"⚠️ Дисциплины не найдены в БД: {', '.join(sorted(missing_disciplines)[:5])}")
    
    seen = set()

    for idx, row in df.iterrows():
        try:
            date = pd.to_datetime(row['Дата']).date()
            pair = int(row['Пара'])
            group_name = str(row['Группа']).strip()
            discipline_name = str(row['Дисциплина']).strip()
            teacher_name = str(row['Преподаватель']).strip()
            
            # Дубли в файле
            key = (date, pair, group_name)
            if key in seen:
                skipped += 1
                continue
            seen.add(key)
            
            group = groups_cache.get(group_name)
            discipline = disciplines_cache.get(discipline_name)
            teacher_parts = teacher_name.split()
            teacher = teachers_cache.get(teacher_parts[0]) if teacher_parts else None
            
            # Собираем причину ошибки
            missing = []
            if not group:
                missing.append(f"группа '{group_name}' не найдена")
            if not discipline:
                missing.append(f"дисциплина '{discipline_name}' не найдена")
            if not teacher:
                missing.append(f"преподаватель '{teacher_name}' не найден")
            
            if missing:
                error_detail = f"Строка {idx+2}: {', '.join(missing)}"
                error_details.append(error_detail)
                errors += 1
                continue
            
            # Конфликт в БД
            existing = Lesson.objects.filter(date=date, pair_number=pair, group=group).first()
            if existing:
                conflict_msg = f"{date} | пара {pair} | {group.name} — уже '{existing.discipline}'"
                conflicts.append(conflict_msg)
                skipped += 1
                continue
            
            lesson = Lesson.objects.create(
                date=date,
                pair_number=pair,
                group=group,
                discipline=discipline,
                teacher=teacher,
                topic=str(row.get('Тема', ''))
            )
            
            students = group.students.all()
            if students.exists():
                grades = [Grade(student=s, lesson=lesson) for s in students]
                Grade.objects.bulk_create(grades, ignore_conflicts=True)
            
            created += 1
            
        except Exception as e:
            error_detail = f"Строка {idx+2}: ошибка обработки — {str(e)}"
            error_details.append(error_detail)
            errors += 1

    show_import_result(request, created, skipped, errors, conflicts, error_details)

def import_students_data(request, df):
    created = 0
    skipped = 0
    errors = 0
    error_details = []
    
    institution = Institution.objects.first()
    if not institution:
        messages.error(request, "❌ Нет учебного заведения.")
        return
    
    group_names = df['Группа'].dropna().unique()
    groups_cache = {
        g.name: g for g in Group.objects.filter(name__in=group_names, institution=institution)
    }
    
    missing_groups = set(group_names) - set(groups_cache.keys())
    if missing_groups:
        messages.warning(request, f"⚠️ Группы не найдены: {', '.join(sorted(missing_groups)[:5])}")
    
    for idx, row in df.iterrows():
        try:
            last_name = str(row['Фамилия']).strip()
            first_name = str(row['Имя']).strip()
            middle_name = str(row.get('Отчество', '')).strip()
            group_name = str(row['Группа']).strip()
            
            if not last_name or not first_name:
                error_details.append(f"Строка {idx+2}: пустые Фамилия или Имя")
                errors += 1
                continue
            
            group = groups_cache.get(group_name)
            if not group:
                error_details.append(f"Строка {idx+2}: группа '{group_name}' не найдена")
                errors += 1
                continue
            
            student_id = str(row.get('Номер студенческого', '')).strip()
            if not student_id:
                import random
                student_id = f"ST{random.randint(10000, 99999)}"
            
            # Проверка уникальности
            if Student.objects.filter(student_id=student_id).exists():
                error_details.append(f"Строка {idx+2}: студенческий '{student_id}' уже существует")
                skipped += 1
                continue
                
            if Student.objects.filter(last_name=last_name, first_name=first_name, group=group).exists():
                skipped += 1
                continue
            
            Student.objects.create(
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                group=group,
                student_id=student_id
            )
            created += 1
                
        except Exception as e:
            error_details.append(f"Строка {idx+2}: {str(e)}")
            errors += 1

    show_import_result(request, created, skipped, errors, error_details=error_details)

def import_teachers_data(request, df):
    created = 0
    skipped = 0
    errors = 0
    error_details = []
    
    institution = Institution.objects.first()
    if not institution:
        messages.error(request, "❌ Нет учебного заведения.")
        return
    
    for idx, row in df.iterrows():
        try:
            last_name = str(row['Фамилия']).strip()
            first_name = str(row['Имя']).strip()
            middle_name = str(row.get('Отчество', '')).strip()
            phone = str(row.get('Телефон', '')).strip()
            
            if not last_name or not first_name:
                error_details.append(f"Строка {idx+2}: пустые Фамилия или Имя")
                errors += 1
                continue
            
            if Teacher.objects.filter(last_name=last_name, first_name=first_name, institution=institution).exists():
                skipped += 1
                continue
            
            # Создаём пользователя
            import random
            username = f"teacher_{last_name.lower()}_{random.randint(10, 99)}"
            
            # Проверка уникальности username
            if User.objects.filter(username=username).exists():
                username = f"{username}_{random.randint(100, 999)}"
            
            user = User.objects.create_user(
                username=username,
                password='Teacher123',
                first_name=first_name,
                last_name=last_name
            )
            
            Teacher.objects.create(
                user=user,
                institution=institution,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                phone=phone
            )
            created += 1
                
        except Exception as e:
            error_details.append(f"Строка {idx+2}: {str(e)}")
            errors += 1

    show_import_result(request, created, skipped, errors, error_details=error_details)

def import_disciplines_data(request, df):
    created = 0
    skipped = 0
    errors = 0
    error_details = []
    
    institution = Institution.objects.first()
    if not institution:
        messages.error(request, "❌ Нет учебного заведения.")
        return
    
    for idx, row in df.iterrows():
        try:
            name = str(row['Название']).strip()
            code = str(row.get('Код', '')).strip()
            
            if not name:
                error_details.append(f"Строка {idx+2}: пустое название дисциплины")
                errors += 1
                continue
            
            if Discipline.objects.filter(name=name, institution=institution).exists():
                skipped += 1
                continue
            
            Discipline.objects.create(name=name, code=code, institution=institution)
            created += 1
            
        except Exception as e:
            error_details.append(f"Строка {idx+2}: {str(e)}")
            errors += 1

    show_import_result(request, created, skipped, errors, error_details=error_details)

def import_groups_data(request, df):
    created = 0
    skipped = 0
    errors = 0
    error_details = []
    
    institution = Institution.objects.first()
    if not institution:
        messages.error(request, "❌ Нет учебного заведения.")
        return
    
    for idx, row in df.iterrows():
        try:
            name = str(row['Название']).strip()
            specialty = str(row.get('Специальность', 'Не указана')).strip()
            start_year = int(row['Год начала']) if 'Год начала' in row else 2025
            
            if not name:
                error_details.append(f"Строка {idx+2}: пустое название группы")
                errors += 1
                continue
            
            if Group.objects.filter(name=name, institution=institution).exists():
                skipped += 1
                continue
            
            Group.objects.create(
                institution=institution,
                name=name,
                specialty=specialty,
                start_year=start_year
            )
            created += 1
            
        except Exception as e:
            error_details.append(f"Строка {idx+2}: {str(e)}")
            errors += 1

    show_import_result(request, created, skipped, errors, error_details=error_details)

def show_import_result(request, created, skipped=0, errors=0, conflicts=None, error_details=None):
    """Показывает результат импорта с детализацией"""
    if created:
        messages.success(request, f"✅ Успешно импортировано: {created} записей")
    
    if skipped:
        messages.warning(request, f"⚠️ Пропущено (дубликаты): {skipped} записей")
    
    if conflicts:
        conflict_msg = "⚠️ Конфликты расписания:\n"
        conflict_msg += "\n".join(f"• {c}" for c in conflicts[:10])
        if len(conflicts) > 10:
            conflict_msg += f"\n... и ещё {len(conflicts) - 10}"
        messages.warning(request, conflict_msg)
    
    if errors:
        error_msg = f"❌ Ошибок при импорте: {errors} записей"
        if error_details:
            error_msg += "\n\nПервые ошибки:\n"
            error_msg += "\n".join(f"• {e}" for e in error_details[:5])
            if len(error_details) > 5:
                error_msg += f"\n... и ещё {len(error_details) - 5}"
        messages.error(request, error_msg)
    
    if not created and not errors and not skipped:
        messages.info(request, "ℹ️ Ничего не импортировано — все данные уже существуют")

def download_template(request, template_type):
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    templates = {
        'schedule': {
            'title': 'Расписание',
            'headers': ['Дата', 'Пара', 'Группа', 'Дисциплина', 'Преподаватель', 'Тема'],
            'example': ['2026-04-18', 1, 'ИС-21', 'Математика', 'Иванов Иван Иванович', 'Пределы']
        },
        'students': {
            'title': 'Студенты',
            'headers': ['Фамилия', 'Имя', 'Отчество', 'Группа', 'Номер студенческого'],
            'example': ['Иванов', 'Иван', 'Иванович', 'ИС-21', 'ST-2025-001']
        },
        'teachers': {
            'title': 'Преподаватели',
            'headers': ['Фамилия', 'Имя', 'Отчество', 'Телефон'],
            'example': ['Петров', 'Пётр', 'Петрович', '+79991234567']
        },
        'disciplines': {
            'title': 'Дисциплины',
            'headers': ['Название', 'Код'],
            'example': ['Математика', 'MATH101']
        },
        'groups': {
            'title': 'Группы',
            'headers': ['Название', 'Специальность', 'Год начала'],
            'example': ['ИС-21', 'Информационные системы', 2025]
        },
    }
    
    config = templates.get(template_type, templates['schedule'])
    ws.title = config['title']
    ws.append(config['headers'])
    ws.append(config['example'])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=template_{template_type}.xlsx'
    wb.save(response)
    return response
@login_required
def export_semester_report(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Проверка прав (только завуч/админ)
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "Доступ запрещён")
        return redirect('teacher_dashboard')
    
    # Получаем teacher (если есть) или берём инфу из User
    try:
        teacher = Teacher.objects.get(user=request.user)
        institution = teacher.institution
    except Teacher.DoesNotExist:
        # Суперюзер без привязки к Teacher — берём первое учреждение
        institution = Institution.objects.first()
        if not institution:
            messages.error(request, "Нет учебного заведения. Создайте учреждение.")
            return redirect('admin_dashboard')
        teacher = None
    
    # Получаем параметры
    group_id = request.GET.get('group_id')
    discipline_id = request.GET.get('discipline_id')
    period_type = request.GET.get('period_type')
    file_format = request.GET.get('format', 'xlsx')
    
    if not group_id or not period_type:
        messages.error(request, "Выберите группу и тип периода")
        return redirect('admin_dashboard')
    
    # Проверяем группу
    group = get_object_or_404(Group, id=group_id)
    
    # Если завуч — проверяем, что группа его учреждения
    if teacher and group.institution != institution:
        messages.error(request, "Эта группа не вашего учебного заведения")
        return redirect('admin_dashboard')
    
    # Определяем период
    if period_type == 'semester':
        semester = int(request.GET.get('semester', 1))
        if semester == 1:
            date_filter = Q(date__month__in=[9, 10, 11, 12])
            period_name = f"1 семестр"
        else:
            date_filter = Q(date__month__in=[1, 2, 3, 4, 5, 6])
            period_name = f"2 семестр"
    
    elif period_type == 'dates':
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if not date_from or not date_to:
            messages.error(request, "Выберите обе даты")
            return redirect('admin_dashboard')
        
        date_filter = Q(date__gte=date_from, date__lte=date_to)
        period_name = f"{date_from} — {date_to}"
    
    elif period_type == 'year':
        year_start = int(request.GET.get('academic_year', 2025))
        date_filter = Q(date__gte=f"{year_start}-09-01", date__lte=f"{year_start+1}-06-30")
        period_name = f"{year_start}/{year_start+1} учебный год"
    
    else:
        messages.error(request, "Неверный тип периода")
        return redirect('admin_dashboard')
    
    # Фильтруем занятия
    lessons = Lesson.objects.filter(
        date_filter,
        group=group
    ).select_related('discipline', 'teacher')
    
    # Фильтр по дисциплине
    if discipline_id:
        lessons = lessons.filter(discipline_id=discipline_id)
    
    lessons = lessons.order_by('date', 'pair_number')
    
    if not lessons.exists():
        messages.warning(request, f"Нет занятий для группы {group.name} за {period_name}")
        return redirect('admin_dashboard')
    
    # Создаём Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ведомость {group.name}"
    
    # Стили
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Ведомость успеваемости | {group.name} | {period_name}"
    title_cell.font = Font(bold=True, size=16, color='1a1a2e')
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    # Подзаголовок
    ws.merge_cells('A2:H2')
    sub_cell = ws.cell(row=2, column=1)
    teacher_name = f"{teacher.last_name} {teacher.first_name}" if teacher else request.user.username
    sub_cell.value = f"Завуч: {teacher_name} | Дата выгрузки: {datetime.now().strftime('%d.%m.%Y')}"
    sub_cell.font = Font(size=10, color='666666')
    sub_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25
    
    row = 4
    
    # Получаем студентов группы
    students = group.students.all().order_by('last_name', 'first_name')
    
    # Получаем все даты занятий (для заголовков колонок)
    lesson_dates = []
    for lesson in lessons:
        date_str = lesson.date.strftime('%d.%m')
        pair_str = f"{date_str} ({lesson.pair_number}п)"
        lesson_dates.append({
            'id': lesson.id,
            'label': pair_str,
            'discipline': lesson.discipline.name,
            'date': lesson.date,
            'pair': lesson.pair_number
        })
    
    # Заголовки таблицы
    headers = ['№', 'Студент']
    for ld in lesson_dates:
        headers.append(ld['label'])
    headers.append('Ср. балл')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    row += 1
    
    # Заполняем оценки
    for idx, student in enumerate(students, 1):
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=f"{student.last_name} {student.first_name}").border = thin_border
        ws.cell(row=row, column=2).alignment = cell_alignment
        
        total_score = 0
        score_count = 0
        
        for col_offset, ld in enumerate(lesson_dates, 3):
            grade = Grade.objects.filter(
                student=student,
                lesson_id=ld['id']
            ).first()
            
            cell = ws.cell(row=row, column=col_offset)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if grade:
                value = grade.value
                cell.value = value
                
                # Подсветка оценок
                if value == '5':
                    cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                elif value == '4':
                    cell.fill = PatternFill(start_color='d1ecf1', end_color='d1ecf1', fill_type='solid')
                elif value == '3':
                    cell.fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
                elif value == '2':
                    cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                elif value in ('Н', 'н'):
                    cell.fill = PatternFill(start_color='e2e3e5', end_color='e2e3e5', fill_type='solid')
                
                # Считаем средний балл
                if value.isdigit():
                    total_score += int(value)
                    score_count += 1
        
        # Средний балл
        avg_cell = ws.cell(row=row, column=len(headers))
        if score_count > 0:
            avg = round(total_score / score_count, 2)
            avg_cell.value = avg
        else:
            avg_cell.value = '—'
        
        avg_cell.font = Font(bold=True)
        avg_cell.border = thin_border
        avg_cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Итоговая строка
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="Среднее по предметам:").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    
    row += 2
    
    # Легенда
    legend_data = [
        ('5', 'd4edda', 'Отлично'),
        ('4', 'd1ecf1', 'Хорошо'),
        ('3', 'fff3cd', 'Удовлетворительно'),
        ('2', 'f8d7da', 'Неудовлетворительно'),
        ('Н', 'e2e3e5', 'Не был'),
    ]
    
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="Легенда:").font = Font(bold=True)
    row += 1
    
    for value, color, desc in legend_data:
        cell = ws.cell(row=row, column=1, value=value)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=desc).border = thin_border
        row += 1
    
    # Сохраняем
    if file_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=report_{group.name}_{period_name}.csv'
        # CSV сложнее с форматированием, пока отдаём Excel
    else:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=report_{group.name}_{period_name}.xlsx'
    
    wb.save(response)
    return response


    lesson = Lesson.objects.get(id=lesson_id)

    disciplines = Discipline.objects.all()
    groups = Group.objects.all()

    if request.method == "POST":
        lesson.discipline_id = request.POST.get("discipline")
        lesson.group_id = request.POST.get("group")
        lesson.date = request.POST.get("date")
        lesson.topic = request.POST.get("topic")

        lesson.save()

        return redirect("admin_dashboard")  # или куда тебе нужно

    return render(request, "journal/edit_lesson.html", {
        "lesson": lesson,
        "disciplines": disciplines,
        "groups": groups
    })

@login_required
def discipline_list(request):
    """Список дисциплин"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')
    
    if request.user.is_superuser:
        # Суперпользователь видит все дисциплины
        disciplines = Discipline.objects.all()
    else:
        try:
            teacher = Teacher.objects.get(user=request.user)
            disciplines = Discipline.objects.filter(institution=teacher.institution)
        except Teacher.DoesNotExist:
            disciplines = Discipline.objects.none()
            messages.warning(request, 'Профиль преподавателя не найден')
    
    context = {
        'disciplines': disciplines,
    }
    return render(request, 'journal/discipline_list.html', context)

@login_required
def discipline_create(request):
    """Создание новой дисциплины"""
    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('login')

    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code', '')
        
        # Определяем учебное заведение
        if request.user.is_superuser:
            institution_id = request.POST.get('institution')
            if not institution_id:
                messages.error(request, 'Выберите учебное заведение')
                institutions = Institution.objects.all()
                return render(request, 'journal/discipline_create.html', {'institutions': institutions})
            institution = get_object_or_404(Institution, id=institution_id)
        else:
            try:
                teacher = Teacher.objects.get(user=request.user)
                institution = teacher.institution
            except Teacher.DoesNotExist:
                messages.error(request, 'Профиль преподавателя не найден')
                return redirect('discipline_list')
        
        # Проверка на существование такой дисциплины в этом заведении
        if Discipline.objects.filter(name=name, institution=institution).exists():
            messages.error(request, f'Дисциплина "{name}" уже существует в вашем учебном заведении!')
            return redirect('discipline_create')
        
        # Создаём дисциплину
        Discipline.objects.create(
            name=name,
            code=code,
            institution=institution
        )
        messages.success(request, 'Дисциплина успешно создана!')
        return redirect('discipline_list')

    # GET запрос
    context = {}
    if request.user.is_superuser:
        context['institutions'] = Institution.objects.all()
    
    return render(request, 'journal/discipline_create.html', context)

@login_required
def discipline_edit(request, pk):
    discipline = Discipline.objects.get(id=pk)

    if request.method == "POST":
        discipline.name = request.POST.get("name")
        discipline.code = request.POST.get("code")
        discipline.save()
        return redirect('discipline_list')

    return render(request, 'journal/discipline_edit.html', {
        'discipline': discipline
    })
    
@login_required
def discipline_delete(request, pk):
    Discipline.objects.get(id=pk).delete()
    return redirect('discipline_list')

@login_required
def admin_schedule(request, date=None):
    """Расписание для администратора/завуча"""
    # проверка прав
    if not request.user.is_staff:
        return redirect('student_schedule')

    # дата
    if date:
        try:
            current_date = datetime.strptime(date, "%Y-%m-%d").date()
        except:
            current_date = datetime.today().date()
    else:
        current_date = datetime.today().date()

    previous_date = (current_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (current_date + timedelta(days=1)).strftime("%Y-%m-%d")

    # Базовый запрос
    lessons = Lesson.objects.filter(date=current_date)
    
    # Фильтрация по правам пользователя
    if request.user.is_superuser:
        # Суперпользователь видит все уроки
        lessons = lessons.select_related('discipline', 'teacher', 'group', 'group__institution')
    else:
        # Завуч (staff) видит только уроки своего учебного заведения
        try:
            teacher = Teacher.objects.get(user=request.user)
            lessons = lessons.filter(
                group__institution=teacher.institution
            ).select_related('discipline', 'teacher', 'group', 'group__institution')
        except Teacher.DoesNotExist:
            lessons = Lesson.objects.none()
            messages.warning(request, 'Ваш профиль не связан с учебным заведением')
    
    # Сортировка для правильной группировки
    lessons = lessons.order_by('group__name', 'pair_number')
    
    context = {
        'lessons': lessons,
        'date': current_date,
        'previous_date': previous_date,
        'next_date': next_date,
        'is_superuser': request.user.is_superuser,
    }
    return render(request, 'journal/admin_schedule.html', context)
    
@login_required
def delete_lesson(request, lesson_id):
    if not request.user.is_staff:
        return redirect('student_schedule')

    lesson = Lesson.objects.get(id=lesson_id)

    if request.method == "POST":
        lesson.delete()
        return redirect('admin_schedule')


